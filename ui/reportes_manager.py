# ui/reportes_manager.py
"""
Interfaz para la gestión de reportes
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import os
import csv
from database.models import get_db_session, Venta, Pago, Cliente, Nicho
from reports.pdf_generator import PDFGenerator

class ReportesManager:
    def __init__(self, parent, update_status_callback):
        self.parent = parent
        self.update_status = update_status_callback
        self.pdf_generator = PDFGenerator()
        
    def show(self):
        """Mostrar interfaz de gestión de reportes"""
        main_frame = ttk.LabelFrame(self.parent, text="Generación de Reportes", padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Frame de configuración
        config_frame = ttk.LabelFrame(main_frame, text="Configuración del Reporte", padding="15")
        config_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        config_frame.columnconfigure(1, weight=1)
        
        self.create_config_widgets(config_frame)
        
        # Frame de vista previa/resultados
        preview_frame = ttk.LabelFrame(main_frame, text="Vista Previa", padding="10")
        preview_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(0, weight=1)
        
        self.create_preview_widgets(preview_frame)
        
        # Frame de acciones
        actions_frame = ttk.Frame(main_frame)
        actions_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        self.create_action_buttons(actions_frame)
    
    def create_config_widgets(self, parent):
        """Crear widgets de configuración"""
        # Tipo de reporte
        ttk.Label(parent, text="Tipo de Reporte:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.tipo_reporte = tk.StringVar(value="movimientos")
        tipo_combo = ttk.Combobox(parent, textvariable=self.tipo_reporte, width=30, state="readonly")
        tipo_combo['values'] = (
            'movimientos', 'ventas', 'pagos', 'clientes', 'nichos', 
            'saldos_pendientes', 'resumen_financiero'
        )
        tipo_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        tipo_combo.bind('<<ComboboxSelected>>', self.on_tipo_changed)
        
        # Período de tiempo
        ttk.Label(parent, text="Período:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.periodo = tk.StringVar(value="mes_actual")
        periodo_combo = ttk.Combobox(parent, textvariable=self.periodo, width=30, state="readonly")
        periodo_combo['values'] = (
            'hoy', 'ayer', 'semana_actual', 'semana_pasada', 
            'mes_actual', 'mes_pasado', 'trimestre_actual', 
            'año_actual', 'personalizado', 'todo'
        )
        periodo_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        periodo_combo.bind('<<ComboboxSelected>>', self.on_periodo_changed)
        
        # Fechas personalizadas (inicialmente ocultas)
        self.fecha_frame = ttk.Frame(parent)
        self.fecha_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(self.fecha_frame, text="Desde:").grid(row=0, column=0, padx=(0, 5))
        self.fecha_inicio = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(self.fecha_frame, textvariable=self.fecha_inicio, width=12).grid(row=0, column=1)
        
        ttk.Label(self.fecha_frame, text="Hasta:").grid(row=0, column=2, padx=(20, 5))
        self.fecha_fin = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(self.fecha_frame, textvariable=self.fecha_fin, width=12).grid(row=0, column=3)
        
        # Ocultar fechas personalizadas inicialmente
        self.fecha_frame.grid_remove()
        
        # Formato de salida
        ttk.Label(parent, text="Formato:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.formato = tk.StringVar(value="pdf")
        formato_frame = ttk.Frame(parent)
        formato_frame.grid(row=3, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        
        ttk.Radiobutton(formato_frame, text="PDF", variable=self.formato, 
                       value="pdf").pack(side=tk.LEFT)
        ttk.Radiobutton(formato_frame, text="CSV", variable=self.formato, 
                       value="csv").pack(side=tk.LEFT, padx=(20, 0))
        ttk.Radiobutton(formato_frame, text="Excel", variable=self.formato, 
                       value="excel").pack(side=tk.LEFT, padx=(20, 0))
        
        # Filtros adicionales
        ttk.Label(parent, text="Filtros:").grid(row=4, column=0, sticky=(tk.W, tk.N), pady=5)
        filtros_frame = ttk.Frame(parent)
        filtros_frame.grid(row=4, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        self.incluir_anulados = tk.BooleanVar(value=False)
        ttk.Checkbutton(filtros_frame, text="Incluir registros anulados", 
                       variable=self.incluir_anulados).pack(anchor=tk.W)
        
        self.solo_pagados = tk.BooleanVar(value=False)
        ttk.Checkbutton(filtros_frame, text="Solo pagados completamente", 
                       variable=self.solo_pagados).pack(anchor=tk.W)
        
        # Agrupar por
        ttk.Label(parent, text="Agrupar por:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.agrupar_por = tk.StringVar(value="fecha")
        agrupar_combo = ttk.Combobox(parent, textvariable=self.agrupar_por, width=30, state="readonly")
        agrupar_combo['values'] = ('fecha', 'titular', 'nicho', 'metodo_pago', 'tipo_pago', 'seccion')
        agrupar_combo.grid(row=5, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        # Botón para generar vista previa
        ttk.Button(parent, text="Generar Vista Previa", 
                  command=self.generate_preview).grid(row=6, column=0, columnspan=2, pady=20)
    
    def create_preview_widgets(self, parent):
        """Crear widgets de vista previa"""
        # TreeView para mostrar datos
        self.preview_tree = ttk.Treeview(parent, show='headings', height=15)
        self.preview_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=self.preview_tree.yview)
        h_scrollbar = ttk.Scrollbar(parent, orient=tk.HORIZONTAL, command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Frame de estadísticas
        stats_frame = ttk.Frame(parent)
        stats_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        self.stats_label = ttk.Label(stats_frame, text="Seleccione un tipo de reporte y genere la vista previa", 
                                    font=("Arial", 10))
        self.stats_label.pack()
    
    def create_action_buttons(self, parent):
        """Crear botones de acción"""
        ttk.Button(parent, text="Exportar Reporte",
                  command=self.export_report).pack(side=tk.LEFT, padx=5)

        ttk.Button(parent, text="Imprimir",
                  command=self.print_report).pack(side=tk.LEFT, padx=5)

        # Botones comentados - funcionalidad no implementada aún
        # ttk.Button(parent, text="Enviar por Email",
        #           command=self.email_report).pack(side=tk.LEFT, padx=5)

        # ttk.Button(parent, text="Programar Reporte",
        #           command=self.schedule_report).pack(side=tk.LEFT, padx=5)

        ttk.Button(parent, text="Limpiar",
                  command=self.clear_preview).pack(side=tk.RIGHT, padx=5)
    
    def on_tipo_changed(self, event=None):
        """Manejar cambio de tipo de reporte"""
        tipo = self.tipo_reporte.get()
        
        # Ajustar opciones según el tipo
        if tipo in ['clientes', 'nichos']:
            # Para reportes de maestros, el período no es tan relevante
            self.periodo.set('todo')
        elif tipo == 'resumen_financiero':
            # Para resumen financiero, sugerir mes actual
            self.periodo.set('mes_actual')
    
    def on_periodo_changed(self, event=None):
        """Manejar cambio de período"""
        if self.periodo.get() == 'personalizado':
            self.fecha_frame.grid()
        else:
            self.fecha_frame.grid_remove()
    
    def get_date_range(self):
        """Obtener rango de fechas según el período seleccionado"""
        periodo = self.periodo.get()
        today = datetime.now().date()
        
        if periodo == 'hoy':
            return today, today
        elif periodo == 'ayer':
            yesterday = today - timedelta(days=1)
            return yesterday, yesterday
        elif periodo == 'semana_actual':
            start = today - timedelta(days=today.weekday())
            return start, today
        elif periodo == 'semana_pasada':
            start = today - timedelta(days=today.weekday() + 7)
            end = start + timedelta(days=6)
            return start, end
        elif periodo == 'mes_actual':
            start = today.replace(day=1)
            return start, today
        elif periodo == 'mes_pasado':
            last_month = today.replace(day=1) - timedelta(days=1)
            start = last_month.replace(day=1)
            return start, last_month
        elif periodo == 'trimestre_actual':
            quarter_start_month = ((today.month - 1) // 3) * 3 + 1
            start = today.replace(month=quarter_start_month, day=1)
            return start, today
        elif periodo == 'año_actual':
            start = today.replace(month=1, day=1)
            return start, today
        elif periodo == 'personalizado':
            try:
                start = datetime.strptime(self.fecha_inicio.get(), "%Y-%m-%d").date()
                end = datetime.strptime(self.fecha_fin.get(), "%Y-%m-%d").date()
                return start, end
            except ValueError:
                messagebox.showerror("Error", "Formato de fecha inválido. Use YYYY-MM-DD")
                return None, None
        else:  # 'todo'
            return None, None
    
    def generate_preview(self):
        """Generar vista previa del reporte"""
        try:
            tipo = self.tipo_reporte.get()
            fecha_inicio, fecha_fin = self.get_date_range()
            
            if self.periodo.get() == 'personalizado' and (fecha_inicio is None or fecha_fin is None):
                return
            
            # Limpiar vista previa anterior
            self.clear_preview()
            
            # Generar datos según el tipo de reporte
            if tipo == 'movimientos':
                data, columns = self.get_movimientos_data(fecha_inicio, fecha_fin)
            elif tipo == 'ventas':
                data, columns = self.get_ventas_data(fecha_inicio, fecha_fin)
            elif tipo == 'pagos':
                data, columns = self.get_pagos_data(fecha_inicio, fecha_fin)
            elif tipo == 'clientes':
                data, columns = self.get_clientes_data(fecha_inicio, fecha_fin)
            elif tipo == 'nichos':
                data, columns = self.get_nichos_data()
            elif tipo == 'saldos_pendientes':
                data, columns = self.get_saldos_pendientes_data()
            elif tipo == 'resumen_financiero':
                data, columns = self.get_resumen_financiero_data(fecha_inicio, fecha_fin)
            else:
                messagebox.showerror("Error", "Tipo de reporte no implementado")
                return
            
            # Configurar TreeView
            self.preview_tree['columns'] = columns
            for col in columns:
                self.preview_tree.heading(col, text=col.replace('_', ' ').title())
                self.preview_tree.column(col, width=100)
            
            # Agregar datos
            for row in data:
                self.preview_tree.insert('', 'end', values=row)
            
            # Actualizar estadísticas
            self.update_stats(data, tipo)
            
            self.update_status(f"Vista previa generada: {len(data)} registros")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar vista previa: {str(e)}")
    
    def get_movimientos_data(self, fecha_inicio, fecha_fin):
        """Obtener datos de movimientos (ventas y pagos)"""
        db = get_db_session()
        
        movimientos = []
        columns = ['fecha', 'tipo', 'numero', 'cliente', 'concepto', 'monto', 'estado']
        
        # Obtener ventas
        query_ventas = db.query(Venta).join(Cliente)
        if fecha_inicio and fecha_fin:
            query_ventas = query_ventas.filter(
                Venta.fecha_venta.between(
                    datetime.combine(fecha_inicio, datetime.min.time()),
                    datetime.combine(fecha_fin, datetime.max.time())
                )
            )
        
        for venta in query_ventas.all():
            estado = "Pagado" if venta.pagado_completamente else "Pendiente"
            movimientos.append([
                venta.fecha_venta.strftime('%d/%m/%Y'),
                'Venta',
                venta.numero_contrato,
                venta.cliente.nombre_completo,
                f"Venta nicho {venta.nicho.numero}",
                f"${venta.precio_total:,.2f}",
                estado
            ])
        
        # Obtener pagos
        query_pagos = db.query(Pago).join(Venta).join(Cliente)
        if fecha_inicio and fecha_fin:
            query_pagos = query_pagos.filter(
                Pago.fecha_pago.between(
                    datetime.combine(fecha_inicio, datetime.min.time()),
                    datetime.combine(fecha_fin, datetime.max.time())
                )
            )
        
        for pago in query_pagos.all():
            movimientos.append([
                pago.fecha_pago.strftime('%d/%m/%Y'),
                'Pago',
                pago.numero_recibo,
                pago.venta.cliente.nombre_completo,
                pago.concepto,
                f"${pago.monto:,.2f}",
                "Procesado"
            ])
        
        db.close()
        
        # Ordenar por fecha
        movimientos.sort(key=lambda x: datetime.strptime(x[0], '%d/%m/%Y'), reverse=True)
        
        return movimientos, columns
    
    def get_ventas_data(self, fecha_inicio, fecha_fin):
        """Obtener datos de ventas"""
        db = get_db_session()
        
        query = db.query(Venta).join(Cliente).join(Nicho)
        
        if fecha_inicio and fecha_fin:
            query = query.filter(
                Venta.fecha_venta.between(
                    datetime.combine(fecha_inicio, datetime.min.time()),
                    datetime.combine(fecha_fin, datetime.max.time())
                )
            )
        
        if self.solo_pagados.get():
            query = query.filter(Venta.pagado_completamente == True)
        
        ventas = query.order_by(Venta.fecha_venta.desc()).all()
        
        columns = ['fecha', 'contrato', 'cliente', 'nicho', 'precio', 'tipo_pago', 'saldo', 'estado']
        
        data = []
        for venta in ventas:
            estado = "Pagado" if venta.pagado_completamente else "Pendiente"
            data.append([
                venta.fecha_venta.strftime('%d/%m/%Y'),
                venta.numero_contrato,
                venta.cliente.nombre_completo,
                venta.nicho.numero,
                f"${venta.precio_total:,.2f}",
                venta.tipo_pago.title(),
                f"${venta.saldo_restante:,.2f}",
                estado
            ])
        
        db.close()
        return data, columns
    
    def get_pagos_data(self, fecha_inicio, fecha_fin):
        """Obtener datos de pagos"""
        db = get_db_session()
        
        query = db.query(Pago).join(Venta).join(Cliente)
        
        if fecha_inicio and fecha_fin:
            query = query.filter(
                Pago.fecha_pago.between(
                    datetime.combine(fecha_inicio, datetime.min.time()),
                    datetime.combine(fecha_fin, datetime.max.time())
                )
            )
        
        pagos = query.order_by(Pago.fecha_pago.desc()).all()
        
        columns = ['fecha', 'recibo', 'contrato', 'cliente', 'monto', 'metodo', 'concepto']
        
        data = []
        for pago in pagos:
            data.append([
                pago.fecha_pago.strftime('%d/%m/%Y'),
                pago.numero_recibo,
                pago.venta.numero_contrato,
                pago.venta.cliente.nombre_completo,
                f"${pago.monto:,.2f}",
                pago.metodo_pago.title(),
                pago.concepto
            ])
        
        db.close()
        return data, columns
    
    def get_clientes_data(self, fecha_inicio, fecha_fin):
        """Obtener datos de clientes"""
        db = get_db_session()
        
        query = db.query(Cliente)
        
        if fecha_inicio and fecha_fin:
            query = query.filter(
                Cliente.fecha_registro.between(
                    datetime.combine(fecha_inicio, datetime.min.time()),
                    datetime.combine(fecha_fin, datetime.max.time())
                )
            )
        
        clientes = query.order_by(Cliente.fecha_registro.desc()).all()
        
        columns = ['nombre', 'cedula', 'telefono', 'email', 'ventas', 'monto_total', 'fecha_registro']
        
        data = []
        for cliente in clientes:
            total_ventas = len(cliente.ventas)
            monto_total = sum(venta.precio_total for venta in cliente.ventas)
            
            data.append([
                cliente.nombre_completo,
                cliente.cedula,
                cliente.telefono or 'N/A',
                cliente.email or 'N/A',
                str(total_ventas),
                f"${monto_total:,.2f}",
                cliente.fecha_registro.strftime('%d/%m/%Y')
            ])
        
        db.close()
        return data, columns
    
    def get_nichos_data(self):
        """Obtener datos de nichos"""
        db = get_db_session()
        
        nichos = db.query(Nicho).order_by(Nicho.seccion, Nicho.fila, Nicho.columna).all()
        
        columns = ['numero', 'seccion', 'ubicacion', 'precio', 'estado', 'cliente', 'fecha_venta']
        
        data = []
        for nicho in nichos:
            estado = "Disponible" if nicho.disponible else "Vendido"
            cliente = ""
            fecha_venta = ""
            
            if nicho.ventas:
                venta = nicho.ventas[0]  # Última venta
                cliente = venta.cliente.nombre_completo
                fecha_venta = venta.fecha_venta.strftime('%d/%m/%Y')
            
            precio_text = f"${nicho.precio:,.2f}" if nicho.precio is not None else "Sin precio"
            data.append([
                nicho.numero,
                nicho.seccion,
                f"F{nicho.fila}-C{nicho.columna}",
                precio_text,
                estado,
                cliente,
                fecha_venta
            ])
        
        db.close()
        return data, columns
    
    def get_saldos_pendientes_data(self):
        """Obtener datos de saldos pendientes"""
        db = get_db_session()
        
        ventas = db.query(Venta).join(Cliente).filter(
            Venta.pagado_completamente == False
        ).order_by(Venta.saldo_restante.desc()).all()
        
        columns = ['contrato', 'cliente', 'nicho', 'precio_total', 'pagado', 'saldo', 'dias_vencido']
        
        data = []
        for venta in ventas:
            dias_vencido = (datetime.now().date() - venta.fecha_venta.date()).days
            
            data.append([
                venta.numero_contrato,
                venta.cliente.nombre_completo,
                venta.nicho.numero,
                f"${venta.precio_total:,.2f}",
                f"${venta.total_pagado:,.2f}",
                f"${venta.saldo_restante:,.2f}",
                str(dias_vencido)
            ])
        
        db.close()
        return data, columns
    
    def get_resumen_financiero_data(self, fecha_inicio, fecha_fin):
        """Obtener resumen financiero"""
        db = get_db_session()
        
        # Calcular totales
        total_ventas = 0
        total_pagos = 0
        total_mantenimiento = 0
        ventas_count = 0
        pagos_count = 0
        mantenimiento_count = 0

        # Ventas en el período
        query_ventas = db.query(Venta)
        if fecha_inicio and fecha_fin:
            query_ventas = query_ventas.filter(
                Venta.fecha_venta.between(
                    datetime.combine(fecha_inicio, datetime.min.time()),
                    datetime.combine(fecha_fin, datetime.max.time())
                )
            )

        ventas = query_ventas.all()
        for venta in ventas:
            total_ventas += venta.precio_total
            ventas_count += 1

        # Pagos en el período (separando mantenimiento de pagos normales)
        query_pagos = db.query(Pago)
        if fecha_inicio and fecha_fin:
            query_pagos = query_pagos.filter(
                Pago.fecha_pago.between(
                    datetime.combine(fecha_inicio, datetime.min.time()),
                    datetime.combine(fecha_fin, datetime.max.time())
                )
            )

        pagos = query_pagos.all()
        for pago in pagos:
            if pago.concepto == 'Mantenimiento':
                total_mantenimiento += pago.monto
                mantenimiento_count += 1
            else:
                total_pagos += pago.monto
                pagos_count += 1
        
        # Saldos pendientes (total)
        saldo_pendiente = db.query(Venta).filter(
            Venta.pagado_completamente == False
        ).with_entities(db.func.sum(Venta.saldo_restante)).scalar() or 0
        
        db.close()
        
        columns = ['concepto', 'cantidad', 'monto']

        data = [
            ['Ventas del Período', str(ventas_count), f"${total_ventas:,.2f}"],
            ['Pagos del Período (sin mantenimiento)', str(pagos_count), f"${total_pagos:,.2f}"],
            ['Pagos de Mantenimiento', str(mantenimiento_count), f"${total_mantenimiento:,.2f}"],
            ['Saldos Pendientes', '-', f"${saldo_pendiente:,.2f}"],
            ['Diferencia (Pagos - Ventas)', '-', f"${total_pagos - total_ventas:,.2f}"]
        ]
        
        return data, columns
    
    def update_stats(self, data, tipo):
        """Actualizar estadísticas de la vista previa"""
        total_registros = len(data)
        
        if tipo in ['movimientos', 'ventas', 'pagos', 'resumen_financiero']:
            # Calcular totales monetarios
            total_monto = 0
            for row in data:
                monto_str = None
                if tipo == 'movimientos':
                    monto_str = row[5]  # columna monto
                elif tipo == 'ventas':
                    monto_str = row[4]  # columna precio
                elif tipo == 'pagos':
                    monto_str = row[4]  # columna monto
                elif tipo == 'resumen_financiero':
                    monto_str = row[2]  # columna monto
                
                if monto_str and monto_str.startswith('$'):
                    try:
                        monto = float(monto_str.replace('$', '').replace(',', ''))
                        total_monto += monto
                    except ValueError:
                        pass
            
            stats_text = f"Total de registros: {total_registros} | Monto total: ${total_monto:,.2f}"
        else:
            stats_text = f"Total de registros: {total_registros}"
        
        self.stats_label.config(text=stats_text)
    
    def clear_preview(self):
        """Limpiar vista previa"""
        # Limpiar TreeView
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        # Limpiar columnas
        self.preview_tree['columns'] = ()
        
        # Limpiar estadísticas
        self.stats_label.config(text="Seleccione un tipo de reporte y genere la vista previa")
    
    def export_report(self):
        """Exportar reporte"""
        # Verificar que hay datos en la vista previa
        if not self.preview_tree.get_children():
            messagebox.showwarning("Advertencia", "Genere primero la vista previa del reporte")
            return
        
        formato = self.formato.get()
        tipo = self.tipo_reporte.get()
        
        # Generar nombre de archivo por defecto
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"reporte_{tipo}_{timestamp}"
        
        try:
            if formato == 'pdf':
                filename = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    filetypes=[("PDF files", "*.pdf")],
                    initialfile=f"{default_filename}.pdf"
                )
                if filename:
                    self.export_to_pdf(filename)

            elif formato == 'csv':
                filename = filedialog.asksaveasfilename(
                    defaultextension=".csv",
                    filetypes=[("CSV files", "*.csv")],
                    initialfile=f"{default_filename}.csv"
                )
                if filename:
                    self.export_to_csv(filename)

            elif formato == 'excel':
                filename = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    initialfile=f"{default_filename}.xlsx"
                )
                if filename:
                    self.export_to_excel(filename)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar reporte: {str(e)}")
    
    def export_to_pdf(self, filename):
        """Exportar a PDF"""
        # Obtener datos de la vista previa
        columns = self.preview_tree['columns']
        data = []
        
        for item in self.preview_tree.get_children():
            values = self.preview_tree.item(item)['values']
            data.append(dict(zip(columns, values)))
        
        # Obtener rango de fechas
        fecha_inicio, fecha_fin = self.get_date_range()
        
        # Generar PDF usando el generador de reportes
        pdf_path = self.pdf_generator.generar_reporte_movimientos(
            data, fecha_inicio or datetime.now().date(), 
            fecha_fin or datetime.now().date(), filename
        )
        
        messagebox.showinfo("Éxito", f"Reporte exportado exitosamente a:\n{pdf_path}")
        
        # Preguntar si desea abrir el archivo
        if messagebox.askyesno("Abrir Archivo", "¿Desea abrir el archivo generado?"):
            os.startfile(pdf_path)  # Windows
    
    def export_to_csv(self, filename):
        """Exportar a CSV"""
        columns = self.preview_tree['columns']
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Escribir encabezados
            writer.writerow([col.replace('_', ' ').title() for col in columns])
            
            # Escribir datos
            for item in self.preview_tree.get_children():
                values = self.preview_tree.item(item)['values']
                writer.writerow(values)
        
        messagebox.showinfo("Éxito", f"Reporte exportado exitosamente a:\n{filename}")
    
    def export_to_excel(self, filename):
        """Exportar a Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"
            
            columns = self.preview_tree['columns']
            
            # Escribir encabezados con formato
            for col_num, column in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num, value=column.replace('_', ' ').title())
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Escribir datos
            for row_num, item in enumerate(self.preview_tree.get_children(), 2):
                values = self.preview_tree.item(item)['values']
                for col_num, value in enumerate(values, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            messagebox.showinfo("Éxito", f"Reporte exportado exitosamente a:\n{filename}")
            
        except ImportError:
            messagebox.showerror("Error", "La biblioteca openpyxl no está instalada.\n"
                               "Instale con: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel: {str(e)}")
    
    def print_report(self):
        """Imprimir reporte"""
        # Verificar que hay datos
        if not self.preview_tree.get_children():
            messagebox.showwarning("Advertencia", "Genere primero la vista previa del reporte")
            return
        
        # Crear PDF temporal para imprimir
        try:
            temp_filename = f"temp_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            self.export_to_pdf(temp_filename)
            
            # Abrir PDF para imprimir
            os.startfile(temp_filename)  # Windows
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al imprimir reporte: {str(e)}")
    
    def email_report(self):
        """Enviar reporte por email"""
        messagebox.showinfo("Funcionalidad", 
                           "La funcionalidad de envío por email estará disponible en una próxima versión.\n"
                           "Por ahora, puede exportar el reporte y enviarlo manualmente.")
    
    def schedule_report(self):
        """Programar reporte automático"""
        messagebox.showinfo("Funcionalidad", 
                           "La funcionalidad de programación de reportes estará disponible en una próxima versión.")